# -*- coding: utf-8 -*-
import json
import os
import shutil
import tempfile
import unittest
from io import BytesIO
from unittest import mock

from openpyxl import load_workbook

from mocktc_app import app as app_module


class MockTcMaintenanceTestCase(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.mkdtemp(prefix="mocktc-maint-")
        self.data_dir = os.path.join(self.tempdir, "data")
        self.fixture_dir = os.path.join(self.tempdir, "fixtures")
        os.makedirs(self.data_dir)
        os.makedirs(self.fixture_dir)
        self.fixture_name = "editable.json"
        self.fixture_path = os.path.join(self.fixture_dir, self.fixture_name)
        self.rows = [
            {"bom_level": 0, "parent_uid": "", "child_uid": "ROOT", "part_id": "ROOT-1", "part_name": "根总成", "revision_id": "A", "quantity": "1", "unit": "EA"},
            {"bom_level": 1, "parent_uid": "ROOT", "child_uid": "CHILD-A", "part_id": "PART-A", "part_name": "子件A", "revision_id": "A", "quantity": "2", "unit": "EA"},
            {"bom_level": 2, "parent_uid": "CHILD-A", "child_uid": "CHILD-B", "part_id": "PART-B", "part_name": "子件B", "revision_id": "A", "quantity": "3", "unit": "EA"},
        ]
        with open(self.fixture_path, "w", encoding="utf-8") as handle:
            json.dump(self.rows, handle, ensure_ascii=False)
        self.originals = (app_module.DATA_DIR, app_module.DB_PATH, app_module.FIXTURE_DIR)
        app_module.DATA_DIR = self.data_dir
        app_module.DB_PATH = os.path.join(self.data_dir, "mocktc.db")
        app_module.FIXTURE_DIR = self.fixture_dir
        self.env = mock.patch.dict(os.environ, {"MOCKTC_ADMIN_TOKEN": "test-admin-token"}, clear=False)
        self.env.start()
        self.app = app_module.create_app()
        self.app.config.update(TESTING=True)
        self.client = self.app.test_client()
        self.headers = {"X-MockTC-Admin-Token": "test-admin-token"}

    def tearDown(self):
        self.env.stop()
        app_module.DATA_DIR, app_module.DB_PATH, app_module.FIXTURE_DIR = self.originals
        shutil.rmtree(self.tempdir)

    def body(self, response):
        return json.loads(response.get_data(as_text=True))

    def test_fixture_ui_lists_and_searches_rows(self):
        overview = self.client.get("/data")
        self.assertEqual(overview.status_code, 200)
        self.assertIn(b"editable.json", overview.data)
        page = self.client.get("/data/fixture/editable.json?q=PART-B")
        self.assertEqual(page.status_code, 200)
        self.assertIn(b"PART-B", page.data)
        self.assertIn("当前筛选 1 行".encode("utf-8"), page.data)
        self.assertIn(b"fixture-editor", page.data)

    def test_fixture_patch_requires_admin_and_is_immediately_visible_with_backup(self):
        path = "/tc/v1/fixtures/editable.json/rows/CHILD-A"
        denied = self.client.patch(path, json={"quantity": "7"})
        self.assertEqual(denied.status_code, 403)
        changed = self.client.patch(path, headers=self.headers, json={"quantity": "7", "part_name": "已修改子件"})
        self.assertEqual(changed.status_code, 200, changed.get_data(as_text=True))
        payload = self.body(changed)["data"]
        self.assertTrue(payload["backup"].endswith(".bak"))
        readback = self.body(self.client.get("/tc/v1/fixtures/editable.json/query?child_uid=CHILD-A"))
        self.assertEqual(readback["data"]["items"][0]["quantity"], "7")
        self.assertEqual(readback["data"]["items"][0]["part_name"], "已修改子件")
        self.assertTrue(os.path.isfile(os.path.join(self.fixture_dir, ".history", payload["backup"])))

    def test_fixture_add_and_cascade_delete_preserve_valid_tree(self):
        created = self.client.post(
            "/tc/v1/fixtures/editable.json/rows", headers=self.headers,
            json={"parent_uid": "CHILD-B", "part_id": "PART-C", "part_name": "子件C", "quantity": "4", "unit": "PC"},
        )
        self.assertEqual(created.status_code, 201, created.get_data(as_text=True))
        child_uid = self.body(created)["data"]["row"]["child_uid"]
        readback = self.body(self.client.get("/tc/v1/fixtures/editable.json/query?child_uid=" + child_uid))
        self.assertEqual(readback["data"]["items"][0]["bom_level"], 3)
        conflict = self.client.delete("/tc/v1/fixtures/editable.json/rows/CHILD-A", headers=self.headers)
        self.assertEqual(conflict.status_code, 409)
        deleted = self.client.delete("/tc/v1/fixtures/editable.json/rows/CHILD-A?cascade=1", headers=self.headers)
        self.assertEqual(deleted.status_code, 200)
        remaining = self.body(self.client.get("/tc/v1/fixtures/editable.json"))["data"]["items"]
        self.assertEqual([row["child_uid"] for row in remaining], ["ROOT"])

    def test_standard_bom_patch_create_and_delete(self):
        changed = self.client.patch(
            "/tc/v1/bomlines/bl-p1000-1", headers=self.headers,
            json={"quantity": 2.5, "unit": "PC", "notes": "测试修改"},
        )
        self.assertEqual(changed.status_code, 200, changed.get_data(as_text=True))
        line = self.body(changed)["data"]
        self.assertEqual(line["quantity"], 2.5)
        self.assertEqual(line["unit"], "PC")
        created = self.client.post(
            "/tc/v1/items/item-p1000/bomlines", headers=self.headers,
            json={"child_item_uid": "item-m1105", "position": "0090", "sequence": 90, "quantity": 3, "unit": "EA", "notes": "新增测试"},
        )
        self.assertEqual(created.status_code, 201, created.get_data(as_text=True))
        line_uid = self.body(created)["data"]["uid"]
        deleted = self.client.delete("/tc/v1/bomlines/" + line_uid, headers=self.headers)
        self.assertEqual(deleted.status_code, 200)
        self.assertEqual(self.client.get("/tc/v1/bomlines/" + line_uid).status_code, 404)

    def test_one_click_excel_and_fixture_json_downloads(self):
        page = self.client.get("/data")
        self.assertIn("一键下载全部数据".encode("utf-8"), page.data)
        exported = self.client.get("/tc/v1/export.xlsx")
        self.assertEqual(exported.status_code, 200)
        self.assertIn("attachment", exported.headers.get("Content-Disposition", ""))
        self.assertIn("mocktc-all-data-", exported.headers.get("Content-Disposition", ""))
        workbook = load_workbook(BytesIO(exported.data), read_only=True, data_only=True)
        self.assertEqual(
            workbook.sheetnames,
            ["导出说明", "物料清单", "标准BOM", "外部BOM_ROOT-1"],
        )
        self.assertGreaterEqual(workbook["物料清单"].max_row, 22)
        self.assertEqual(workbook["外部BOM_ROOT-1"].max_row, 4)
        downloaded = self.client.get("/tc/v1/fixtures/editable.json/download")
        self.assertEqual(downloaded.status_code, 200)
        self.assertEqual(downloaded.content_type, "application/json")
        self.assertIn('attachment; filename="editable.json"', downloaded.headers["Content-Disposition"])
        self.assertEqual(json.loads(downloaded.data), self.rows)


if __name__ == "__main__":
    unittest.main()
